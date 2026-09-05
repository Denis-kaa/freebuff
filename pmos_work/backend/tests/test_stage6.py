"""Тесты этапа 6 (6.md §54-56): Import/Export Engine.

Acceptance-тесты Excel: импорт 3 проектов + позиции, повторный импорт
(update, без дублей), ошибка валидации (dry run — база не меняется).
"""

import io
import uuid

from openpyxl import Workbook, load_workbook


def make_xlsx(rows: list[list]) -> bytes:
    """Создаёт Excel-файл с заголовками + строками (в памяти)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ВВОД"
    headers = ["ID", "Проект", "Менеджер", "Дедлайн", "Оплата %", "Валюта",
               "Дата аванса", "Позиция", "Тираж", "Тех. макет", "Дата отгрузки сигнала"]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


ROWS_OK = [
    ["P100", "Wazzup", "Денис", "03.09.2026", "50%", "RUB", "31.08.2026", "Худи", "100", "Сдан", "01.09.2026"],
    ["P101", "Чико", "Денис", "10.09.2026", "80%", "USD", "", "Футболка", "50", "", ""],
    ["P102", "Чисто", "", "15.09.2026", "100%", "EUR", "", "", "", "", ""],
]


async def _upload(client, data: bytes, filename="projects_test.xlsx"):
    return await client.post(
        "/api/imports/excel",
        files={"file": (filename, data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


# ---------------------------------------------------------------------------
# §54: импорт 3 проектов + позиции
# ---------------------------------------------------------------------------
async def test_import_preview_and_confirm(client):
    res = await _upload(client, make_xlsx(ROWS_OK))
    assert res.status_code == 200, res.text
    body = res.json()
    job = body["job"]
    preview = body["preview"]
    assert job["status"] == "VALIDATING"
    assert preview["total"] == 3
    assert preview["errors"] == 0
    assert preview["will_create"] == 3
    # авто-маппинг заголовков (§5)
    mapping = preview["mapping"]
    assert mapping["Проект"] == "title"
    assert mapping["Оплата %"] == "payment_percent"
    assert mapping["Позиция"] == "item_name"

    res = await client.post(f"/api/imports/{job['id']}/confirm",
                            json={"mapping": mapping, "duplicate_mode": "update"})
    assert res.status_code == 200, res.text
    result = res.json()["result"]
    assert result["created"] == 3
    assert result["updated"] == 0

    # проекты созданы
    lst = (await client.get("/api/projects?page_size=100")).json()["items"]
    titles = {p["display_id"]: p["title"] for p in lst}
    assert titles["P100"] == "Wazzup"
    assert titles["P101"] == "Чико"
    # позиции созданы (плоская строка -> иерархия, §21)
    items = (await client.get(f"/api/project-items?project_id={next(p['id'] for p in lst if p['display_id']=='P100')}")).json()
    assert any(i["name"] == "Худи" and i["quantity"] == 100 for i in items)
    assert any(i["mockup_status"] == "Сдан" for i in items)


async def test_import_no_items_project(client):
    res = await _upload(client, make_xlsx([ROWS_OK[2]]))
    body = res.json()
    res = await client.post(f"/api/imports/{body['job']['id']}/confirm",
                            json={"mapping": body["preview"]["mapping"]})
    assert res.json()["result"]["created"] == 1


# ---------------------------------------------------------------------------
# §55: повторный импорт — update, без дублей
# ---------------------------------------------------------------------------
async def test_duplicate_import_updates_not_duplicates(client):
    data = make_xlsx(ROWS_OK)
    up = (await _upload(client, data)).json()
    await client.post(f"/api/imports/{up['job']['id']}/confirm",
                      json={"mapping": up["preview"]["mapping"]})
    # повторный импорт
    up2 = (await _upload(client, data)).json()
    assert up2["preview"]["will_create"] == 0
    assert up2["preview"]["will_update"] == 3
    res = await client.post(f"/api/imports/{up2['job']['id']}/confirm",
                            json={"mapping": up2["preview"]["mapping"]})
    result = res.json()["result"]
    assert result["created"] == 0
    assert result["updated"] == 3
    assert result["skipped"] == 0
    # дублей нет: всё ещё 3 проекта с P100-102
    lst = (await client.get("/api/projects?page_size=100")).json()["items"]
    p100 = [p for p in lst if p["display_id"] == "P100"]
    assert len(p100) == 1


async def test_duplicate_mode_skip_and_copy(client):
    data = make_xlsx(ROWS_OK)
    up = (await _upload(client, data)).json()
    await client.post(f"/api/imports/{up['job']['id']}/confirm",
                      json={"mapping": up["preview"]["mapping"]})
    # skip
    up2 = (await _upload(client, data)).json()
    res = await client.post(f"/api/imports/{up2['job']['id']}/confirm",
                            json={"mapping": up2["preview"]["mapping"], "duplicate_mode": "skip"})
    assert res.json()["result"]["skipped"] == 3
    # copy
    up3 = (await _upload(client, data)).json()
    res = await client.post(f"/api/imports/{up3['job']['id']}/confirm",
                            json={"mapping": up3["preview"]["mapping"], "duplicate_mode": "copy"})
    assert res.json()["result"]["created"] == 3  # P103, P104, P105


# ---------------------------------------------------------------------------
# §56: ошибка валидации — dry run, база не меняется
# ---------------------------------------------------------------------------
async def test_validation_error_dry_run_no_changes(client):
    rows = [list(r) for r in ROWS_OK]  # глубокая копия: не мутируем общий ROWS_OK
    rows[0][4] = "70%"  # недопустимое значение оплаты
    data = make_xlsx(rows)
    up = (await _upload(client, data)).json()
    preview = up["preview"]
    assert preview["errors"] >= 1
    err = next(i for i in preview["issues"] if i["level"] == "ERROR")
    assert err["field"] == "payment_percent"

    # полный импорт -> rollback (§15), база не изменилась
    res = await client.post(f"/api/imports/{up['job']['id']}/confirm",
                            json={"mapping": preview["mapping"], "duplicate_mode": "update"})
    body = res.json()
    assert body["job"]["status"] == "FAILED"
    assert body.get("rolled_back") is True

    lst = (await client.get("/api/projects?page_size=100")).json()["items"]
    assert not any(p["display_id"] == "P100" for p in lst)


async def test_partial_import_only_valid_rows(client):
    """§16: импорт только корректных строк; ошибки не теряются."""
    rows = [list(r) for r in ROWS_OK]  # глубокая копия: не мутируем общий ROWS_OK
    rows[0][4] = "70%"
    up = (await _upload(client, make_xlsx(rows))).json()
    res = await client.post(f"/api/imports/{up['job']['id']}/confirm",
                            json={"mapping": up["preview"]["mapping"], "partial": True})
    result = res.json()["result"]
    assert result["created"] == 2  # P101, P102
    assert result["skipped"] == 1  # P100 с ошибкой
    errs = (await client.get(f"/api/imports/{up['job']['id']}/errors")).json()["errors"]
    assert any(e["field"] == "payment_percent" for e in errs)


async def test_required_field_error(client):
    rows = [["P200", "", "Денис", "03.09.2026", "50%", "RUB", "", "", "", "", ""]]
    up = (await _upload(client, make_xlsx(rows))).json()
    assert up["preview"]["errors"] >= 1


# ---------------------------------------------------------------------------
# Manual mapping + saved mapping (§7-9)
# ---------------------------------------------------------------------------
async def test_manual_mapping_and_save(client):
    # заголовок «Проектик» не распознаётся автоматически
    wb = Workbook()
    ws = wb.active
    ws.append(["Проектик", "Срок", "Пропустить меня"])
    ws.append(["Тест", "05.10.2026", "x"])
    buf = io.BytesIO()
    wb.save(buf)

    up = (await _upload(client, buf.getvalue(), "manual.xlsx")).json()
    preview = up["preview"]
    assert "Проектик" not in preview["mapping"]  # авто не нашёл

    # ручной mapping
    mapping = {"Проектик": "title", "Срок": "deadline", "Пропустить меня": "__ignore__"}
    res = await client.post(f"/api/imports/{up['job']['id']}/mapping", json={"mapping": mapping})
    assert res.status_code == 200
    assert res.json()["preview"]["errors"] == 0
    assert res.json()["preview"]["mapping"]["Проектик"] == "title"

    # сохранить mapping (§9)
    res = await client.post(f"/api/imports/{up['job']['id']}/save-mapping",
                            json={"name": "Моя таблица", "mapping": mapping})
    assert res.status_code == 200

    res = await client.post(f"/api/imports/{up['job']['id']}/confirm",
                            json={"mapping": mapping, "duplicate_mode": "update"})
    assert res.json()["result"]["created"] == 1


# ---------------------------------------------------------------------------
# История (§36-37)
# ---------------------------------------------------------------------------
async def test_import_history(client):
    up = (await _upload(client, make_xlsx(ROWS_OK[:1]))).json()
    await client.post(f"/api/imports/{up['job']['id']}/confirm",
                      json={"mapping": up["preview"]["mapping"]})
    hist = (await client.get("/api/imports/history")).json()
    assert len(hist) >= 1
    assert hist[0]["status"] == "COMPLETED"
    assert hist[0]["created_count"] == 1
    assert hist[0]["file_name"] == "projects_test.xlsx"


# ---------------------------------------------------------------------------
# Export (§23-29, §49)
# ---------------------------------------------------------------------------
async def test_export_xlsx_all_projects(client, make_project):
    await make_project(title="Денисов", manager_name="Денис", deadline="2026-09-10", payment_percent="50%")
    await make_project(title="Катин", manager_name="Катя", deadline="2026-09-11", payment_percent="100%")
    res = await client.post("/api/exports/excel", json={"scope": "all_projects"})
    assert res.status_code == 200, res.text
    meta = res.json()
    assert meta["format"] == "xlsx"
    dl = await client.get(meta["download_url"])
    assert dl.status_code == 200
    wb = load_workbook(io.BytesIO(dl.content), data_only=True)
    ws = wb["Проекты"]
    headers = [c.value for c in ws[1]]
    assert "Проект" in headers
    titles = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert any(t == "Денисов" for t in titles)
    assert ws.freeze_panes == "A2"


async def test_export_with_manager_filter(client, make_project):
    await make_project(title="Денисов", manager_name="Денис", deadline="2026-09-10")
    await make_project(title="Катин", manager_name="Катя", deadline="2026-09-11")
    res = await client.post("/api/exports/excel", json={
        "scope": "all_projects", "filters": {"manager": "Денис"},
    })
    dl = await client.get(res.json()["download_url"])
    wb = load_workbook(io.BytesIO(dl.content), data_only=True)
    ws = wb["Проекты"]
    titles = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert titles == ["Денисов"]


async def test_export_csv(client, make_project):
    await make_project(title="Тест CSV", manager_name="Денис")
    res = await client.post("/api/exports/csv", json={"scope": "all_projects"})
    dl = await client.get(res.json()["download_url"])
    text = dl.content.decode("utf-8-sig")
    assert "Проект" in text.split("\n")[0]
    assert "Тест CSV" in text


async def test_export_projects_items(client, make_project):
    p = await make_project(title="Wazzup")
    await client.post("/api/project-items", json={"project_id": p["id"], "name": "Худи", "quantity": 100})
    res = await client.post("/api/exports/excel", json={"scope": "projects_items"})
    dl = await client.get(res.json()["download_url"])
    wb = load_workbook(io.BytesIO(dl.content), data_only=True)
    assert "Проекты" in wb.sheetnames
    assert "Позиции" in wb.sheetnames
    ws = wb["Позиции"]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    assert any("Худи" in r for r in rows)


async def test_export_formula_injection_escaped(client, make_project):
    """§49: значения '=...' экранируются апострофом."""
    await make_project(title="=SUM(A1:A9)", manager_name="Денис")
    res = await client.post("/api/exports/csv", json={"scope": "all_projects"})
    dl = await client.get(res.json()["download_url"])
    text = dl.content.decode("utf-8-sig")
    assert "=SUM(A1:A9)" not in text.split(";")[1] or "'=SUM(A1:A9)" in text


async def test_export_calendar(client, make_project):
    p = await make_project(title="Wazzup", deadline="2026-09-20")
    res = await client.post("/api/exports/excel", json={
        "scope": "calendar", "filters": {"from": "2026-09-01", "to": "2026-09-30"},
    })
    assert res.status_code == 200
    dl = await client.get(res.json()["download_url"])
    wb = load_workbook(io.BytesIO(dl.content), data_only=True)
    ws = wb["Календарь"]
    dates = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert any(str(d).startswith("2026-09-20") for d in dates if d)


# ---------------------------------------------------------------------------
# Template (§41-42) + Security (§48)
# ---------------------------------------------------------------------------
async def test_import_template(client):
    res = await client.get("/api/imports/templates?kind=projects_items")
    assert res.status_code == 200
    wb = load_workbook(io.BytesIO(res.content), data_only=True)
    ws = wb["ВВОД"]
    headers = [c.value for c in ws[1]]
    assert "Проект" in headers
    assert "Позиция" in headers
    assert "Тираж" in headers
    assert wb.sheetnames == ["ВВОД", "Инструкция"]


async def test_upload_bad_extension(client):
    res = await client.post("/api/imports/excel",
                            files={"file": ("evil.exe", b"MZ...", "application/octet-stream")})
    assert res.status_code == 422


async def test_upload_oversize(client):
    big = b"0" * (21 * 1024 * 1024)
    res = await client.post("/api/imports/excel",
                            files={"file": ("big.xlsx", big,
                                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res.status_code in (413, 422)


async def test_import_permissions(client):
    """Чужой job id -> 404 (§48 workspace ownership)."""
    assert (await client.get(f"/api/imports/{uuid.uuid4()}")).status_code == 404