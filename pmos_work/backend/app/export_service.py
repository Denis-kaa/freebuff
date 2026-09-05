"""Export Engine (6.md §23-29, §41-42, §49).

PostgreSQL -> Flatten/Mapping -> .xlsx / .csv. Excel — внешний интерфейс,
не источник истины. Форматирование: frozen header, autofilter, ширины,
даты/проценты/валюта (§27). Защита от formula injection (§49).
"""
import csv
import os
import re
import uuid
from datetime import date, datetime
from typing import Any, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from .models import Project, ProjectItem, Task

EXPORT_DIR = "/var/www/pm_os/exports"

# Колонки проектов для экспорта (6.md §25)
PROJECT_EXPORT_COLUMNS: list[dict] = [
    {"field": "display_id", "label": "ID"},
    {"field": "title", "label": "Проект"},
    {"field": "client_legal_name", "label": "Юр. лицо"},
    {"field": "manager_name", "label": "Менеджер"},
    {"field": "stage", "label": "Этап"},
    {"field": "deadline", "label": "Дедлайн"},
    {"field": "payment_percent", "label": "Оплата %"},
    {"field": "currency", "label": "Валюта"},
    {"field": "advance_date", "label": "Дата аванса"},
    {"field": "final_payment_date", "label": "Дата доплаты"},
    {"field": "delivery_address", "label": "Адрес доставки"},
    {"field": "delivery_paid", "label": "Доставка оплачена"},
    {"field": "next_action", "label": "Следующее действие"},
    {"field": "next_action_date", "label": "Дата след. действия"},
    {"field": "risk_level", "label": "Риск"},
    {"field": "risk_reason", "label": "Причина риска"},
    {"field": "comment", "label": "Комментарий"},
]

ITEM_EXPORT_COLUMNS: list[dict] = [
    {"field": "display_id", "label": "Проект ID"},
    {"field": "project_title", "label": "Проект"},
    {"field": "name", "label": "Позиция"},
    {"field": "quantity", "label": "Тираж"},
    {"field": "tech_specs", "label": "ТЗ"},
    {"field": "mockup_status", "label": "Тех. макет"},
    {"field": "signal_status", "label": "Сигнал"},
    {"field": "signal_shipping_date", "label": "Дата отгрузки сигнала"},
    {"field": "signal_feedback", "label": "ОС по сигналу"},
    {"field": "batch_status", "label": "Тираж"},
    {"field": "factory", "label": "Фабрика"},
]

TASK_EXPORT_COLUMNS: list[dict] = [
    {"field": "display_id", "label": "Проект ID"},
    {"field": "project_title", "label": "Проект"},
    {"field": "title", "label": "Задача"},
    {"field": "status", "label": "Статус"},
    {"field": "priority", "label": "Приоритет"},
    {"field": "assignee_name", "label": "Ответственный"},
    {"field": "due_date", "label": "Срок"},
]

# Legacy-формат (6.md §29, §43): старые колонки -> новые поля
LEGACY_EXPORT_COLUMNS: list[dict] = [
    {"field": "display_id", "label": "ID"},
    {"field": "title", "label": "Проект"},
    {"field": "client_legal_name", "label": "Юр. лицо"},
    {"field": "manager_name", "label": "Менеджер"},
    {"field": "stage", "label": "Этап"},
    {"field": "deadline", "label": "Дедлайн"},
    {"field": "payment_percent", "label": "Оплата %"},
    {"field": "currency", "label": "Валюта"},
    {"field": "advance_date", "label": "Дата аванса"},
    {"field": "final_payment_date", "label": "Дата доплаты"},
    {"field": "delivery_address", "label": "Адрес доставки"},
    {"field": "next_action", "label": "Следующее действие"},
    {"field": "item_names", "label": "Позиции"},
    {"field": "item_quantities", "label": "Тиражи"},
]

FORMULA_CHARS = ("=", "+", "-", "@")


def safe_value(value: Any) -> Any:
    """Formula injection protection (6.md §49)."""
    if isinstance(value, str):
        s = value.strip()
        if s and s[0] in FORMULA_CHARS:
            return "'" + s
    return value


def _fmt_date(v: Any) -> Any:
    if isinstance(v, (date, datetime)):
        return v
    return v


def _fmt_percent(v: Any) -> Any:
    return v if v is None else str(v)


def _fmt_currency(v: Any) -> Any:
    return v


def value_of(project: Project, field: str) -> Any:
    if field == "item_names":
        return "; ".join(i.name for i in project.items)
    if field == "item_quantities":
        return "; ".join(str(i.quantity) for i in project.items if i.quantity)
    return getattr(project, field, None)


def _columns_by_fields(columns: list[dict], fields: Optional[list[str]]) -> list[dict]:
    if not fields:
        return columns
    wanted = set(fields)
    return [c for c in columns if c["field"] in wanted]


# ---------------------------------------------------------------------------
# Сбор данных
# ---------------------------------------------------------------------------
async def fetch_projects(db, workspace_id, filters: Optional[dict] = None, include_archived=False):
    stmt = select(Project).where(Project.workspace_id == workspace_id)
    if not include_archived:
        stmt = stmt.where(Project.archived_at.is_(None))
    filters = filters or {}
    manager = filters.get("manager")
    if manager:
        stmt = stmt.where(Project.manager_name == manager)
    risk = filters.get("risk_level")
    if risk:
        stmt = stmt.where(Project.risk_level == risk)
    stage = filters.get("stage")
    if stage:
        stmt = stmt.where(Project.stage == stage)
    search = filters.get("search")
    if search:
        stmt = stmt.where(Project.title.ilike(f"%{search}%"))
    rows = (await db.execute(stmt)).scalars().all()
    # eager items для item_names/legacy
    from sqlalchemy.orm import selectinload

    if any(f in ("item_names", "item_quantities") for f in (filters.get("columns") or [])) or True:
        stmt2 = select(Project).where(Project.id.in_([p.id for p in rows])).options(selectinload(Project.items))
        rows = (await db.execute(stmt2)).scalars().all()
    return list(rows)


async def fetch_items(db, workspace_id, project_ids: Optional[list[uuid.UUID]] = None):
    stmt = select(ProjectItem).where(ProjectItem.project_id.in_(project_ids)) if project_ids else select(ProjectItem)
    rows = (await db.execute(stmt)).scalars().all()
    return list(rows)


async def fetch_tasks(db, workspace_id, project_ids: Optional[list[uuid.UUID]] = None):
    stmt = select(Task).where(Task.project_id.in_(project_ids)) if project_ids else select(Task)
    rows = (await db.execute(stmt)).scalars().all()
    return list(rows)


# ---------------------------------------------------------------------------
# Запись файлов
# ---------------------------------------------------------------------------
def _write_xlsx(path: str, sheets: list[dict]) -> str:
    """sheets: [{name, headers: [str], rows: [[..]], col_widths: [int]}] (6.md §27-28)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    thin = Alignment(vertical="center")

    for s in sheets:
        ws = wb.create_sheet(title=s["name"][:31])
        ws.append(s["headers"])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = thin
        for row in s["rows"]:
            ws.append([safe_value(v) for v in row])
        # формат дат
        date_cols = s.get("date_cols", [])
        for r in ws.iter_rows(min_row=2, min_col=1, max_col=len(s["headers"])):
            for cell in r:
                if cell.column in date_cols and isinstance(cell.value, (date, datetime)):
                    cell.number_format = "DD.MM.YYYY"
        ws.freeze_panes = "A2"          # frozen header (§27)
        ws.auto_filter.ref = ws.dimensions  # autofilter
        for i, w in enumerate(s.get("col_widths") or [], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(path)
    return path


def _write_csv(path: str, headers: list[str], rows: list[list]) -> str:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(headers)
        for row in rows:
            writer.writerow([safe_value(v) for v in row])
    return path


def _default_widths(labels: list[str]) -> list[int]:
    return [max(12, min(len(l) * 2 + 4, 34)) for l in labels]


# ---------------------------------------------------------------------------
# Основной экспорт
# ---------------------------------------------------------------------------
FINANCE_EXPORT_FIELDS = {"payment_percent", "currency", "advance_date", "final_payment_date"}


def _strip_finance_columns(col_defs: list[dict]) -> list[dict]:
    return [c for c in col_defs if c["field"] not in FINANCE_EXPORT_FIELDS]


def build_export(
    path: str,
    *,
    scope: str,  # current_view|all_projects|projects_items|tasks|calendar|legacy
    projects: list,
    items: Optional[list] = None,
    tasks: Optional[list] = None,
    calendar_events: Optional[list] = None,
    columns: Optional[list[str]] = None,
    fmt: str = "xlsx",  # xlsx|csv
    project_map: Optional[dict] = None,
    can_read_finance: bool = True,
) -> str:
    """Формирует файл экспорта на диске (6.md §23-28).

    can_read_finance=False — финансовые поля исключаются из выгрузки даже
    если пользователь экспортирует всю таблицу (RBAC §47).
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    project_map = project_map or {}

    if scope in ("tasks",):
        col_defs = TASK_EXPORT_COLUMNS
        headers = [c["label"] for c in col_defs]
        rows = []
        for t in tasks or []:
            p = project_map.get(t.project_id)
            rows.append([
                p.display_id if p else "", p.title if p else "", t.title,
                t.status, t.priority or "", t.assignee_name or "", _fmt_date(t.due_date),
            ])
        if fmt == "csv":
            return _write_csv(path, headers, rows)
        return _write_xlsx(path, [{"name": "Задачи", "headers": headers, "rows": rows,
                                   "col_widths": _default_widths(headers), "date_cols": [7]}])

    if scope == "calendar":
        headers = ["Дата", "Тип", "Событие", "Проект", "Статус"]
        rows = []
        for e in calendar_events or []:
            md = e.get("metadata") or {}
            rows.append([
                e["start_at"][:10], e["type"], e["title"],
                md.get("project_display_id") or "", e.get("status") or "",
            ])
        if fmt == "csv":
            return _write_csv(path, headers, rows)
        return _write_xlsx(path, [{"name": "Календарь", "headers": headers, "rows": rows,
                                   "col_widths": _default_widths(headers), "date_cols": [1]}])

    # проектные scope
    if scope == "legacy":
        col_defs = LEGACY_EXPORT_COLUMNS
    else:
        col_defs = _columns_by_fields(PROJECT_EXPORT_COLUMNS, columns)
    if not can_read_finance:
        col_defs = _strip_finance_columns(col_defs)
    headers = [c["label"] for c in col_defs]
    proj_rows = []
    for p in projects:
        proj_rows.append([value_of(p, c["field"]) for c in col_defs])

    date_cols = [i + 1 for i, c in enumerate(col_defs) if c["field"].endswith("_date") or c["field"] == "deadline"]

    if scope == "projects_items" and items is not None:
        item_defs = _columns_by_fields(ITEM_EXPORT_COLUMNS, None)
        item_headers = [c["label"] for c in item_defs]
        item_rows = []
        for it in items:
            p = project_map.get(it.project_id)
            item_rows.append([
                p.display_id if p else "", p.title if p else "", it.name,
                it.quantity, it.tech_specs or "", it.mockup_status or "",
                it.signal_status or "", _fmt_date(it.signal_shipping_date),
                it.signal_feedback or "", it.batch_status or "", it.factory or "",
            ])
        if fmt == "csv":
            _write_csv(path, headers, proj_rows)
            _write_csv(path.replace(".csv", "_items.csv"), item_headers, item_rows)
            return path
        return _write_xlsx(path, [
            {"name": "Проекты", "headers": headers, "rows": proj_rows,
             "col_widths": _default_widths(headers), "date_cols": date_cols},
            {"name": "Позиции", "headers": item_headers, "rows": item_rows,
             "col_widths": _default_widths(item_headers), "date_cols": [8]},
        ])

    if fmt == "csv":
        return _write_csv(path, headers, proj_rows)
    return _write_xlsx(path, [{"name": "Проекты", "headers": headers, "rows": proj_rows,
                               "col_widths": _default_widths(headers), "date_cols": date_cols}])


# ---------------------------------------------------------------------------
# Шаблон импорта (6.md §41-42)
# ---------------------------------------------------------------------------
def build_import_template(path: str, kind: str = "projects") -> str:
    """PROJECT_IMPORT_TEMPLATE.xlsx: колонки, примеры, выпадающие списки, инструкция."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "ВВОД"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")

    if kind == "projects_items":
        headers = ["ID", "Проект", "Юр. лицо", "Менеджер", "Этап", "Дедлайн", "Оплата %", "Валюта",
                   "Дата аванса", "Дата доплаты", "Адрес доставки", "Следующее действие",
                   "Позиция", "Тираж", "Тех. макет", "Сигнал", "Дата отгрузки сигнала", "ОС по сигналу", "Фабрика"]
        example = ["P100", "Wazzup", "ООО Тест", "Денис", "Сигнал", "03.09.2026", "50%", "RUB",
                   "31.08.2026", "05.09.2026", "Москва", "Получить ОС", "Худи", "100", "Сдан", "Отгружен",
                   "01.09.2026", "Ожидается", "Фабрика А"]
        ws.column_dimensions["A"].width = 10
        for i in range(2, 20):
            ws.column_dimensions[get_column_letter(i)].width = 18
    else:  # projects
        headers = ["ID", "Проект", "Юр. лицо", "Менеджер", "Этап", "Дедлайн", "Оплата %", "Валюта",
                   "Дата аванса", "Дата доплаты", "Адрес доставки", "Следующее действие"]
        example = ["P100", "Wazzup", "ООО Тест", "Денис", "Сигнал", "03.09.2026", "50%", "RUB",
                   "31.08.2026", "05.09.2026", "Москва", "Получить ОС"]
        ws.column_dimensions["A"].width = 10
        for i in range(2, 13):
            ws.column_dimensions[get_column_letter(i)].width = 18

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.append(example)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    dv = DataValidation(type="list", formula1='"0%,50%,80%,100%"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"G2:G1000")

    cur = DataValidation(type="list", formula1='"RUB,USD,EUR,USDT"', allow_blank=True)
    ws.add_data_validation(cur)
    cur.add(f"H2:H1000")

    instr = wb.create_sheet("Инструкция")
    lines = [
        ["ИМПОРТ ПРОЕКТОВ", ""],
        ["", ""],
        ["1. Заполните строки, начиная со строки 2 (строка 1 — заголовки).", ""],
        ["2. ID (P100) — можно оставить пустым, система назначит сама.", ""],
        ["3. Если ID уже существует — проект будет обновлён (или выберите «Пропустить»/«Копия» при импорте).", ""],
        ["4. Дата — в формате ДД.ММ.ГГГГ.", ""],
        ["5. Оплата % — одно из: 0%, 50%, 80%, 100% (выпадающий список).", ""],
        ["6. Валюта — RUB, USD, EUR, USDT (выпадающий список).", ""],
        ["7. Позиции: заполняйте «Позиция»/«Тираж» — по одной позиции на строку проекта.", ""],
        ["8. Импорт выполняется через кнопку «Импорт» на странице Проекты.", ""],
    ]
    for row in lines:
        instr.append(row)
    instr.column_dimensions["A"].width = 90
    wb.save(path)
    return path


async def build_calendar_export(db, workspace_id, from_date, to_date) -> list[dict]:
    from .calendar_service import service

    from datetime import datetime, time, timezone as _tz

    frm = datetime.combine(from_date, time.min, tzinfo=_tz.utc)
    to = datetime.combine(to_date, time.max, tzinfo=_tz.utc)
    return await service.get_events(db, workspace_id, frm, to)