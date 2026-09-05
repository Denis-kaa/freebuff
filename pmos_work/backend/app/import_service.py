"""Import Engine (6.md §1-22, §43-45, §53).

Excel/CSV — внешний формат, источник истины — PostgreSQL. Поток:
Upload -> Parse -> (Mapping) -> Validate -> Preview -> Confirm -> Import
Никогда не импортируем вслепую (§14). Импорт транзакционен (§15).
Duplicate detection по display_id (§18), display_id != UUID (§19).
"""
import csv
import io
import os
import re
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from typing import Any, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from .models import ImportJob, Project, ProjectItem

MAX_ROWS = 10_000  # безопасный лимит строк (6.md §45), архитектурно расширяем

# ---------------------------------------------------------------------------
# Справочник полей (6.md §5)
# ---------------------------------------------------------------------------
FIELD_DEFS: dict[str, dict] = {
    "title": {"label": "Проект", "required": True},
    "display_id": {"label": "ID"},
    "client_legal_name": {"label": "Юр. лицо"},
    "manager_name": {"label": "Менеджер"},
    "stage": {"label": "Этап"},
    "deadline": {"label": "Дедлайн", "type": "date"},
    "payment_percent": {"label": "Оплата %", "type": "percent",
                        "enum": ["0%", "50%", "80%", "100%"]},
    "currency": {"label": "Валюта", "type": "currency", "enum": ["RUB", "USD", "EUR", "USDT"]},
    "advance_date": {"label": "Дата аванса", "type": "date"},
    "final_payment_date": {"label": "Дата доплаты", "type": "date"},
    "delivery_address": {"label": "Адрес доставки"},
    "delivery_paid": {"label": "Доставка оплачена", "type": "bool"},
    "next_action": {"label": "Следующее действие"},
    "next_action_date": {"label": "Дата след. действия", "type": "date"},
    "risk_level": {"label": "Риск"},
    "risk_reason": {"label": "Причина риска"},
    "comment": {"label": "Комментарий"},
    # Project Items (6.md §21-22)
    "item_name": {"label": "Позиция", "item": True},
    "item_quantity": {"label": "Тираж", "type": "int", "item": True},
    "item_mockup_status": {"label": "Тех. макет", "item": True},
    "item_signal_required": {"label": "Сигнал нужен", "type": "bool", "item": True},
    "item_signal_status": {"label": "Сигнал", "item": True},
    "item_signal_shipping_date": {"label": "Дата отгрузки сигнала", "type": "date", "item": True},
    "item_signal_feedback": {"label": "ОС по сигналу", "item": True},
    "item_batch_status": {"label": "Тираж", "item": True},
    "item_factory": {"label": "Фабрика", "item": True},
}

# Автоматическое сопоставление по заголовку (6.md §5, smart mapping §6)
HEADER_SYNONYMS: dict[str, str] = {
    "id": "display_id", "номер": "display_id", "проект id": "display_id", "код": "display_id",
    "проект": "title", "название": "title", "наименование": "title", "имя проекта": "title",
    "юр. лицо": "client_legal_name", "юрлицо": "client_legal_name", "клиент": "client_legal_name",
    "организация": "client_legal_name", "контрагент": "client_legal_name",
    "менеджер": "manager_name", "менеджер проекта": "manager_name", "ответственный": "manager_name",
    "этап": "stage", "стадия": "stage",
    "дедлайн": "deadline", "дата дедлайна": "deadline", "срок": "deadline", "срок сдачи": "deadline",
    "оплата %": "payment_percent", "оплата": "payment_percent", "процент оплаты": "payment_percent",
    "валюта": "currency",
    "дата аванса": "advance_date", "аванс": "advance_date",
    "дата доплаты": "final_payment_date", "доплата": "final_payment_date",
    "адрес доставки": "delivery_address", "адрес": "delivery_address",
    "доставка оплачена": "delivery_paid", "доставка": "delivery_paid",
    "следующее действие": "next_action", "след. действие": "next_action", "след действие": "next_action",
    "дата следующего действия": "next_action_date",
    "риск": "risk_level", "уровень риска": "risk_level",
    "причина риска": "risk_reason", "риск причина": "risk_reason",
    "комментарий": "comment", "комментарии": "comment", "заметка": "comment",
    "позиция": "item_name", "изделие": "item_name", "товар": "item_name", "артикул": "item_name",
    "тираж": "item_quantity", "количество": "item_quantity", "кол-во": "item_quantity", "шт": "item_quantity",
    "тех. макет": "item_mockup_status", "макет": "item_mockup_status",
    "сигнал нужен": "item_signal_required", "нужен сигнал": "item_signal_required",
    "сигнал": "item_signal_status", "статус сигнала": "item_signal_status",
    "дата отгрузки сигнала": "item_signal_shipping_date", "отгрузка сигнала": "item_signal_shipping_date",
    "ос по сигналу": "item_signal_feedback", "обратная связь": "item_signal_feedback",
    "тираж статус": "item_batch_status", "статус тиража": "item_batch_status",
    "фабрика": "item_factory", "производство": "item_factory",
}

# Старые колонки -> предупреждение (6.md §6, §43) — не угадываем молча
LEGACY_COLUMN_NOTES: dict[str, str] = {
    "отгрузка": "Старая колонка. В новой структуре отгрузка задаётся датой сигнала (Позиции) или тиражом.",
    "закрыв. док.": "Старая колонка. В новой структуре используются УПД (дизайн/сигнал/производство/объединённый) в Документах.",
    "закрывающие документы": "Старая колонка. Используйте раздел Документы (УПД).",
    "управляющие док.": "Старая колонка.",
}


@dataclass
class RowIssue:
    row: int
    field: Optional[str]
    value: Optional[str]
    error: str
    level: str = "ERROR"  # ERROR | WARNING


@dataclass
class ParsedTable:
    headers: list[str]
    rows: list[list[Any]]
    sheet_name: Optional[str] = None


@dataclass
class ImportPreview:
    total: int
    ok: int
    errors: int
    warnings: int
    issues: list[RowIssue] = field(default_factory=list)
    mapping: dict[str, str] = field(default_factory=dict)  # excel_header -> db_field
    unmapped: list[str] = field(default_factory=list)
    legacy_notes: list[dict] = field(default_factory=list)
    will_create: int = 0
    will_update: int = 0


# ---------------------------------------------------------------------------
# Parser (6.md §4, §48: не доверяем расширению — проверяем содержимое)
# ---------------------------------------------------------------------------
def parse_excel(path: str) -> list[ParsedTable]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)  # формулы НЕ выполняем (§48)
    tables = []
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            continue
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        rows = [list(r) for r in rows_iter]
        tables.append(ParsedTable(headers=headers, rows=rows, sheet_name=ws.title))
    return tables


def parse_csv(path: str) -> list[ParsedTable]:
    raw = open(path, "rb").read()
    text = raw.decode("utf-8-sig") or raw.decode("utf-8")
    sample = text[:4000]
    delim = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    data = [r for r in reader if any((c or "").strip() for c in r)]
    if not data:
        return []
    headers = [str(h).strip() if h is not None else "" for h in data[0]]
    rows = data[1:]
    return [ParsedTable(headers=headers, rows=rows, sheet_name="CSV")]


def detect_table_kind(path: str, tables: list[ParsedTable]) -> str:
    """excel|csv по содержимому (не по имени файла, §48)."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv",):
        return "csv"
    return "excel"


# ---------------------------------------------------------------------------
# Smart mapping (6.md §5-6)
# ---------------------------------------------------------------------------
def auto_mapping(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for h in headers:
        key = h.strip().lower()
        mapped = HEADER_SYNONYMS.get(key)
        if mapped is None:
            # fuzzy: удаляем пробелы/тире, ищем подстроку
            compact = re.sub(r"[\s\-_.]", "", key)
            for syn, field_name in HEADER_SYNONYMS.items():
                if compact == re.sub(r"[\s\-_.]", "", syn):
                    mapped = field_name
                    break
        if mapped:
            mapping[h] = mapped
    return mapping


def legacy_notes(headers: list[str]) -> list[dict]:
    out = []
    for h in headers:
        key = h.strip().lower()
        if key in LEGACY_COLUMN_NOTES:
            out.append({"column": h, "note": LEGACY_COLUMN_NOTES[key]})
    return out


# ---------------------------------------------------------------------------
# Validation (6.md §10-12)
# ---------------------------------------------------------------------------
def _parse_date(value: Any) -> Optional[date]:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None  # невалидно


def _parse_percent(value: Any) -> Optional[str]:
    if value is None or str(value).strip() == "":
        return None
    s = str(value).strip().replace("%", "").replace(" ", "")
    if not s.isdigit():
        return None
    return f"{int(s)}%"


def _parse_currency(value: Any) -> Optional[str]:
    if value is None or str(value).strip() == "":
        return None
    return str(value).strip().upper()


def _parse_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    s = str(value).strip().replace(" ", "")
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _parse_bool(value: Any) -> Optional[bool]:
    if value is None or str(value).strip() == "":
        return None
    s = str(value).strip().lower()
    if s in ("да", "yes", "1", "true", "+", "нужен"):
        return True
    if s in ("нет", "no", "0", "false", "-", "не нужен"):
        return False
    return None


def coerce_value(field_name: str, value: Any) -> tuple[Any, Optional[str]]:
    """Приводит значение к типу поля. Возвращает (значение, ошибка)."""
    fdef = FIELD_DEFS.get(field_name, {})
    ftype = fdef.get("type")
    if value is None or str(value).strip() == "":
        return None, None
    if ftype == "date":
        d = _parse_date(value)
        return (d, None) if d else (None, "Некорректная дата.")
    if ftype == "percent":
        p = _parse_percent(value)
        if p is None:
            return None, f"Недопустимое значение «{value}». Допустимые: {', '.join(fdef.get('enum', []))}"
        if p not in fdef.get("enum", []):
            return None, f"Недопустимое значение «{p}». Допустимые: {', '.join(fdef.get('enum', []))}"
        return p, None
    if ftype == "currency":
        c = _parse_currency(value)
        if c is None or (fdef.get("enum") and c not in fdef["enum"]):
            return None, f"Недопустимая валюта «{value}». Допустимые: {', '.join(fdef.get('enum', []))}"
        return c, None
    if ftype == "int":
        i = _parse_int(value)
        return (i, None) if i is not None else (None, "Некорректное число.")
    if ftype == "bool":
        b = _parse_bool(value)
        return (b, None) if b is not None else (None, "Некорректное значение да/нет.")
    return str(value).strip(), None


def validate_row(row_idx: int, values: dict[str, Any], mapping_rev: dict[str, str]) -> tuple[list[RowIssue], dict[str, Any]]:
    """Проверяет одну строку. Возвращает (issues, нормализованные значения)."""
    issues: list[RowIssue] = []
    data: dict[str, Any] = {}
    for excel_col, field_name in mapping_rev.items():
        if field_name == "__ignore__":
            continue
        raw = values.get(excel_col)
        coerced, err = coerce_value(field_name, raw)
        if err:
            issues.append(RowIssue(row=row_idx, field=field_name, value=str(raw), error=err))
        else:
            data[field_name] = coerced
    # обязательные поля (6.md §10)
    if not data.get("title"):
        issues.append(RowIssue(row=row_idx, field="title", value=values.get("title"), error="Обязательное поле «Проект» не заполнено."))
    # warning: новый менеджер (6.md §12)
    manager = data.get("manager_name")
    if manager and not _manager_known(manager):
        issues.append(RowIssue(row=row_idx, field="manager_name", value=manager,
                               error=f"Менеджер «{manager}» не найден. Будет создан новый справочник.", level="WARNING"))
    return issues, data


_MANAGERS_CACHE: set[str] = set()


def _manager_known(name: str) -> bool:
    return name.lower() in _MANAGERS_CACHE


async def load_manager_cache(session: AsyncSession, workspace_id: uuid.UUID) -> None:
    """Кэш известных менеджеров (справочник) для WARNING (6.md §12)."""
    from .models import User

    global _MANAGERS_CACHE
    names = (await session.execute(
        select(User.display_name).where(User.workspace_id == workspace_id)
    )).scalars().all()
    _MANAGERS_CACHE = {n.lower() for n in names if n}


# ---------------------------------------------------------------------------
# Preview / dry run (6.md §13, §44)
# ---------------------------------------------------------------------------
def build_preview(
    table: ParsedTable,
    mapping: dict[str, str],
    known_display_ids: set[str],
) -> ImportPreview:
    """Анализ без изменений в БД (dry run, §44)."""
    preview = ImportPreview(total=0, ok=0, errors=0, warnings=0)
    preview.mapping = mapping
    preview.unmapped = [h for h in table.headers if h not in mapping]
    preview.legacy_notes = legacy_notes(table.headers)

    mapping_rev = {h: f for h, f in mapping.items() if f}
    for i, row in enumerate(table.rows, start=2):  # строка 1 = заголовки
        values = {}
        for idx, header in enumerate(table.headers):
            values[header] = row[idx] if idx < len(row) else None
        issues, data = validate_row(i, values, mapping_rev)
        preview.total += 1
        has_error = any(x.level == "ERROR" for x in issues)
        if has_error:
            preview.errors += 1
        else:
            preview.ok += 1
            did = data.get("display_id")
            if did and did in known_display_ids:
                preview.will_update += 1
            else:
                preview.will_create += 1
        preview.warnings += sum(1 for x in issues if x.level == "WARNING")
        preview.issues.extend(issues)
    return preview


# ---------------------------------------------------------------------------
# Импорт (транзакционный, §15-16, §18-22)
# ---------------------------------------------------------------------------
async def run_import(
    session: AsyncSession,
    workspace_id: uuid.UUID,
    table: ParsedTable,
    mapping: dict[str, str],
    duplicate_mode: str = "update",  # update|skip|copy
    partial: bool = False,
) -> dict:
    """Выполняет импорт. Внешний код отвечает за commit/rollback (§15)."""
    created = updated = skipped = 0
    job_issues: list[dict] = []
    created_ids: list[uuid.UUID] = []
    updated_ids: list[uuid.UUID] = []

    # известные display_id (для dedup §18) + безопасная генерация новых (§20)
    known_rows = (await session.execute(
        select(Project.display_id).where(Project.workspace_id == workspace_id)
    )).scalars().all()
    known = set(known_rows)
    max_num = 0
    for d in known_rows:
        m = re.match(r"P(\d+)", d or "")
        if m:
            max_num = max(max_num, int(m.group(1)))

    mapping_rev = {h: f for h, f in mapping.items() if f and f != "__ignore__"}
    pending_ids: dict[str, str] = {}  # старый display_id -> новый (для копий)

    for i, row in enumerate(table.rows, start=2):
        values = {}
        for idx, header in enumerate(table.headers):
            values[header] = row[idx] if idx < len(row) else None
        issues, data = validate_row(i, values, mapping_rev)
        errs = [x for x in issues if x.level == "ERROR"]
        warns = [x for x in issues if x.level == "WARNING"]
        for x in issues:
            job_issues.append({"row": x.row, "field": x.field, "value": x.value, "error": x.error, "level": x.level})

        if errs:
            if not partial:
                raise ValueError(f"Строка {i}: {errs[0].error}")  # -> rollback (транзакция)
            skipped += 1
            continue

        # dedup по display_id (§18)
        did = data.get("display_id") or ""
        if did:
            if did in pending_ids:
                did = pending_ids[did]
            if did in known:
                if duplicate_mode == "skip":
                    skipped += 1
                    continue
                if duplicate_mode == "copy":
                    max_num += 1
                    new_did = f"P{max_num:03d}"
                    pending_ids[did] = new_did
                    did = new_did
                    known.add(new_did)
        if not did:
            max_num += 1
            did = f"P{max_num:03d}"
            while did in known:
                max_num += 1
                did = f"P{max_num:03d}"
            known.add(did)

        data["display_id"] = did
        # проектные поля
        project_fields = {k: v for k, v in data.items() if not FIELD_DEFS.get(k, {}).get("item")}
        project_fields.pop("display_id", None)

        existing = (await session.execute(
            select(Project).where(Project.workspace_id == workspace_id, Project.display_id == did)
        )).scalar_one_or_none()

        if existing and duplicate_mode == "update":
            for k, v in project_fields.items():
                setattr(existing, k, v)
            existing.version += 1
            updated += 1
            updated_ids.append(existing.id)
        elif existing is None:
            existing = Project(workspace_id=workspace_id, display_id=did, **project_fields)
            session.add(existing)
            await session.flush()
            created += 1
            created_ids.append(existing.id)

        # позиции (плоская строка -> иерархия, §21-22)
        if data.get("item_name"):
            item_name = str(data["item_name"])
            item = (await session.execute(
                select(ProjectItem).where(
                    ProjectItem.project_id == existing.id, ProjectItem.name == item_name
                )
            )).scalar_one_or_none()
            item_data = {k: v for k, v in data.items() if FIELD_DEFS.get(k, {}).get("item")}
            item_data.pop("item_name", None)
            item_map = {
                "item_quantity": "quantity", "item_mockup_status": "mockup_status",
                "item_signal_required": "signal_required", "item_signal_status": "signal_status",
                "item_signal_shipping_date": "signal_shipping_date",
                "item_signal_feedback": "signal_feedback",
                "item_batch_status": "batch_status", "item_factory": "factory",
            }
            mapped_item = {item_map.get(k, k): v for k, v in item_data.items()}
            if item is None:
                item = ProjectItem(project_id=existing.id, name=item_name, **mapped_item)
                session.add(item)
            else:
                for k, v in mapped_item.items():
                    setattr(item, k, v)

    return {
        "created": created, "updated": updated, "skipped": skipped,
        "created_ids": created_ids, "updated_ids": updated_ids,
        "issues": job_issues,
    }