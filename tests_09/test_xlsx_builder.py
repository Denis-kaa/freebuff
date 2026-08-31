"""tests_09/test_xlsx_builder.py \u2014 xlsx-builder roundtrip + multi-sheet formula (Task 0, ROADMAP-VV-001).

Coverage:
    1. Create + save empty workbook, reopen, sheetnames intact.
    2. Single-sheet: write cell value, reopen, value matches.
    3. Formula persistence across save/load (formula is preserved as `=...`).
    4. Multi-sheet: formula on sheet 2 references sheet 1 \u2014 persists roundtrip.
    5. Value XOR formula constraint: pass both \u2192 ValueError.
    6. Value XOR formula constraint: pass neither \u2192 ValueError.
    7. cell_formula() helper: distinguishes formula cell from value cell.
    8. data_only=True returns cached values (sanity for reviewer role at Task 2).
    9. Formatting (fmt=bold) roundtrips through openpyxl re-read (Font.bold True).

Atomic-save guarantee (PB-7 mirror): save() writes via tempfile + rename.
We verify by saving twice without cleanup \u2014 second save overwrites cleanly.
"""
from __future__ import annotations

from pathlib import Path

import pytest

from core_02.xlsx_builder import Workbook


def test_create_and_roundtrip_empty(tmp_path: Path):
    """Create blank workbook, save, reload \u2014 sheetnames stay consistent."""
    p = tmp_path / "empty.xlsx"
    wb = Workbook()
    wb.save(p)
    assert p.exists()
    loaded = Workbook.load(p)
    assert isinstance(loaded.sheetnames(), list)
    # openpyxl creates 1 default sheet (\"Sheet\")
    assert len(loaded.sheetnames()) >= 1


def test_single_sheet_value_roundtrip(tmp_path: Path):
    """Write cell value on one sheet, reload, value matches."""
    p = tmp_path / "v.xlsx"
    wb = Workbook()
    wb.sheet("\u0418\u0441\u0442\u043e\u0440\u0438\u044f")
    wb.cell("A1", value="\u041a\u0430\u0442\u0435\u0433\u043e\u0440\u0438\u044f")
    wb.cell("B2", value=42)
    wb.save(p)

    loaded = Workbook.load(p)
    assert "\u0418\u0441\u0442\u043e\u0440\u0438\u044f" in loaded.sheetnames()
    loaded.sheet("\u0418\u0441\u0442\u043e\u0440\u0438\u044f")
    assert loaded.cell_value("A1") == "\u041a\u0430\u0442\u0435\u0433\u043e\u0440\u0438\u044f"
    assert loaded.cell_value("B2") == 42


def test_formula_persisted_across_save_load(tmp_path: Path):
    """Formula `=AVERAGE(B2:B10)` on sheet \u041f\u0440\u043e\u0433\u043d\u043e\u0437 survives roundtrip."""
    p = tmp_path / "f.xlsx"
    wb = Workbook()
    wb.sheet("\u041f\u0440\u043e\u0433\u043d\u043e\u0437")
    wb.cell("C5", formula="AVERAGE(B2:B10)")
    wb.save(p)

    loaded = Workbook.load(p, data_only=False)
    loaded.sheet("\u041f\u0440\u043e\u0433\u043d\u043e\u0437")
    assert loaded.cell_formula("C5") == "=AVERAGE(B2:B10)"
    assert loaded.cell_value("C5") == "=AVERAGE(B2:B10)"  # without data_only


def test_multi_sheet_cross_reference(tmp_path: Path):
    """Sheet 2 formula references Sheet 1 cell \u2014 persists roundtrip."""
    p = tmp_path / "multi.xlsx"
    wb = Workbook()
    wb.sheet("\u0418\u0441\u0442\u043e\u0440\u0438\u044f")
    wb.cell("B2", value=100)
    wb.sheet("\u041f\u0440\u043e\u0433\u043d\u043e\u0437")
    wb.cell("C3", formula="'\u0418\u0441\u0442\u043e\u0440\u0438\u044f'!B2 * 1.15")
    wb.save(p)

    loaded = Workbook.load(p, data_only=False)
    assert "\u0418\u0441\u0442\u043e\u0440\u0438\u044f" in loaded.sheetnames()
    assert "\u041f\u0440\u043e\u0433\u043d\u043e\u0437" in loaded.sheetnames()

    loaded.sheet("\u041f\u0440\u043e\u0433\u043d\u043e\u0437")
    assert loaded.cell_formula("C3") == "='\u0418\u0441\u0442\u043e\u0440\u0438\u044f'!B2 * 1.15"

    loaded.sheet("\u0418\u0441\u0442\u043e\u0440\u0438\u044f")
    assert loaded.cell_value("B2") == 100


def test_cell_value_xor_formula_rejects_both():
    """Passing value=... AND formula=... is a hard error."""
    wb = Workbook()
    wb.sheet("S")
    with pytest.raises(ValueError, match="value XOR formula"):
        wb.cell("A1", value=1, formula="AVERAGE(B2:B10)")


def test_cell_value_xor_formula_rejects_neither():
    """Passing neither value=... nor formula=... is a hard error."""
    wb = Workbook()
    wb.sheet("S")
    with pytest.raises(ValueError, match="must pass either"):
        wb.cell("A1")


def test_cell_formula_helper_distinguishes():
    """cell_formula() returns None for pure-value cells, formula string for formula cells."""
    wb = Workbook()
    wb.sheet("T")
    wb.cell("A1", value=5)
    wb.cell("A2", formula="A1 * 2")
    assert wb.cell_formula("A1") is None
    assert wb.cell_formula("A2") == "=A1 * 2"


def test_data_only_returns_cached_value_after_explicit_xlsx_compute(tmp_path: Path):
    """data_only=True read returns None for formula cells (openpyxl limitation,
    doesn't recompute formulas). Anchors that Task 2 reviewer role must NOT rely
    on this \u2014 must restate formulas in Python (Q1 variant (b))."""
    p = tmp_path / "do.xlsx"
    wb = Workbook()
    wb.sheet("Z")
    wb.cell("A1", value=10)
    wb.cell("A2", formula="A1 * 3")
    wb.save(p)

    loaded_data_only = Workbook.load(p, data_only=True)
    loaded_data_only.sheet("Z")
    # openpyxl with data_only=True reads cached computed value;
    # since openpyxl saved without Excel recalc, cached = None.
    assert loaded_data_only.cell_value("A2") is None
    # But formula mode still recovers the formula:
    loaded_formula = Workbook.load(p, data_only=False)
    loaded_formula.sheet("Z")
    assert loaded_formula.cell_formula("A2") == "=A1 * 3"


def test_format_roundtrip_bold(tmp_path: Path):
    """fmt={'bold': True] persists font.bold through roundtrip."""
    p = tmp_path / "fmt.xlsx"
    wb = Workbook()
    wb.sheet("F")
    wb.cell("A1", value="\u0417\u0430\u0433\u043e\u043b\u043e\u0432\u043e\u043a", fmt={"bold": True})
    wb.save(p)

    import openpyxl

    raw = openpyxl.load_workbook(p, data_only=False)
    cell = raw["F"]["A1"]
    assert cell.font.bold is True
    assert cell.value == "\u0417\u0430\u0433\u043e\u043b\u043e\u0432\u043e\u043a"


def test_atomic_save_overwrites_cleanly(tmp_path: Path):
    """save() called twice produces valid file each time (no leftover tmp)."""
    p = tmp_path / "over.xlsx"
    for ver, val in [("v1", 1), ("v2", 2), ("v3", 3)]:
        wb = Workbook()
        wb.sheet("OV")
        wb.cell("A1", value=f"{ver}={val}")
        wb.save(p)
    # No *.tmp* files in tmp_path
    leftovers = list(tmp_path.glob("*.tmp*"))
    assert leftovers == [], f"Failed cleanup on atomic save: {leftovers}"
    loaded = Workbook.load(p)
    loaded.sheet("OV")
    assert loaded.cell_value("A1") == "v3=3"
