"""
test_excel_eval.py — юнит-тесты независимого Excel-formula evaluator
(BUG-005 fix, promt 64 audit, vkusvill_demo).

Два слоя проверок:
1. Реальные ячейки `model_forecast.xlsx` (order_qty, final_forecast, TOTAL,
   cross-sheet refs, blank separator rows, STDEV.P population семантика).
2. Изолированные формулы через временные workbooks (openpyxl) — для
   edge cases, которых нет в demo (MAX с отрицательным входом, VLOOKUP →
   FormulaError, unknown sheet → FormulaError).
"""
from __future__ import annotations

import json
import sys
***REMOVED***

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEMO_DIR = PROJECT_ROOT / "projects_17" / "vkusvill_demo"

sys.path.insert(0, str(DEMO_DIR))
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook  # noqa: E402

from excel_eval import ExcelEval, FormulaError  # noqa: E402

XLSX = DEMO_DIR / "model_forecast.xlsx"
SNAPSHOT = DEMO_DIR / "model_snapshot.json"
PYTHON_OUT = DEMO_DIR / "forecast_python.json"


@pytest.fixture(scope="module")
def xlsx_path() -> Path:
    if not XLSX.exists():
        raise FileNotFoundError(f"{XLSX***REMOVED*** not found. Run build_model_xlsx.py first.")
    return XLSX


@pytest.fixture(scope="module")
def ev(xlsx_path: Path) -> ExcelEval:
    return ExcelEval(xlsx_path)


@pytest.fixture(scope="module")
def snapshot() -> dict:
    return json.loads(SNAPSHOT.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def python_out() -> dict:
    return json.loads(PYTHON_OUT.read_text(encoding="utf-8"))


@pytest.fixture
def make_temp_eval(tmp_path: Path):
    """Фабрика: строит временный xlsx из {sheet: {addr: value***REMOVED******REMOVED*** и возвращает ExcelEval.

    value может быть числом, строкой или формулой ('=...'). Файл создаётся
    в tmp_path pytest (авто-cleanup после теста).
    """
    def _build(cells: dict[str, dict[str, object***REMOVED******REMOVED***) -> ExcelEval:
        tmp = tmp_path / "wb.xlsx"
        wb = Workbook()
        wb.remove(wb.active)  # удалить дефолтный sheet
        for sheet_name, cell_map in cells.items():
            ws = wb.create_sheet(sheet_name)
            for addr, value in cell_map.items():
                ws[addr***REMOVED*** = value
        wb.save(tmp)
        return ExcelEval(tmp)
    return _build


# ---------------------------------------------------------------------------
# Core: реальные ячейки demo xlsx
# ---------------------------------------------------------------------------

class TestCoreCells:
    @pytest.mark.parametrize(
        "row,expected_order,expected_final",
        [
            (4, 760.8975620814633, 398.60122260945303),   # Молоко (dairy, ×0.92)
            (5, 513.3332560153304, 103.59475608143734),   # Крупа
            (6, 662.7583594842604, 220.24463209169693),   # Напиток
        ***REMOVED***,
    )
    def test_order_and_final(self, ev, row, expected_order, expected_final):
        assert ev.evaluate(f"order!E{row***REMOVED***") == pytest.approx(expected_order, abs=1e-9)
        assert ev.evaluate(f"forecast!G{row***REMOVED***") == pytest.approx(expected_final, abs=1e-9)

    def test_total_row_found_dynamically(self, ev):
        hit = ev.find_formula_cell("order", "E", prefix="=SUM(")
        assert hit is not None
        col, row = hit
        total = ev.evaluate(f"order!{col***REMOVED***{row***REMOVED***")
        assert total == pytest.approx(1936.9891775810543, abs=1e-9)

    def test_total_equals_sum_of_orders(self, ev):
        total = ev.evaluate("order!E8")
        parts = sum(ev.evaluate(f"order!E{r***REMOVED***") for r in (4, 5, 6))
        assert total == pytest.approx(parts, abs=1e-9)


# ---------------------------------------------------------------------------
# Cross-sheet / range semantics (real demo xlsx)
# ---------------------------------------------------------------------------

class TestRanges:
    def test_history_range_average(self, ev):
        avg = ev.evaluate("forecast!C4")
        raw = [ev.evaluate(f"history!D{r***REMOVED***") for r in range(12, 16)***REMOVED***
        assert avg == pytest.approx(sum(raw) / 4, abs=1e-9)

    def test_history_range_stdevp(self, ev):
        import statistics

        raw = [ev.evaluate(f"history!D{r***REMOVED***") for r in range(12, 16)***REMOVED***
        expected = statistics.pstdev(raw) * 1.65 * (2 ** 0.5)
        assert ev.evaluate("forecast!E4") == pytest.approx(expected, abs=1e-9)

    def test_blank_separator_ignored(self, ev):
        avg_groats = ev.evaluate("forecast!C5")
        raw = [ev.evaluate(f"history!D{r***REMOVED***") for r in range(25, 29)***REMOVED***
        assert avg_groats == pytest.approx(sum(raw) / 4, abs=1e-9)


# ---------------------------------------------------------------------------
# Функции / семантика (isolated temp workbooks)
# ---------------------------------------------------------------------------

class TestFunctions:
    def test_if_positive_branch(self, make_temp_eval):
        # IF(7<4,0.5,1.0) → 1.0 (условие ложно)
        ev = make_temp_eval({"S": {"A1": "=IF(7<4,0.5,1.0)"***REMOVED******REMOVED***)
        assert ev.evaluate("S!A1") == 1.0

    def test_if_negative_branch(self, make_temp_eval):
        # IF(2<7,0.5,1.0) → 0.5 (условие истинно)
        ev = make_temp_eval({"S": {"A1": "=IF(2<7,0.5,1.0)"***REMOVED******REMOVED***)
        assert ev.evaluate("S!A1") == 0.5

    def test_max_zero_clamp_negative(self, make_temp_eval):
        # MAX(0,-5) → 0 (Excel-семантика: отрицательный вход зажимается)
        ev = make_temp_eval({"S": {"A1": "=MAX(0,-5)"***REMOVED******REMOVED***)
        assert ev.evaluate("S!A1") == 0.0

    def test_max_positive(self, make_temp_eval):
        ev = make_temp_eval({"S": {"A1": "=MAX(0,3,7)"***REMOVED******REMOVED***)
        assert ev.evaluate("S!A1") == 7.0

    def test_sum_range_with_blank(self, make_temp_eval):
        # SUM(A1:A5) с пустой ячейкой A3 → пропускается
        ev = make_temp_eval({"S": {"A1": 1, "A2": 2, "A4": 4, "A5": 5,
                                    "B1": "=SUM(A1:A5)"***REMOVED******REMOVED***)
        assert ev.evaluate("S!B1") == 12.0

    def test_sqrt(self, make_temp_eval):
        ev = make_temp_eval({"S": {"A1": "=SQRT(16)"***REMOVED******REMOVED***)
        assert ev.evaluate("S!A1") == 4.0

    def test_stdevp_is_population(self, ev):
        import statistics

        raw = [ev.evaluate(f"history!D{r***REMOVED***") for r in range(12, 16)***REMOVED***
        safety = ev.evaluate("forecast!E4")
        assert safety == pytest.approx(statistics.pstdev(raw) * 1.65 * (2 ** 0.5), abs=1e-9)

    def test_unsupported_function_raises(self, make_temp_eval):
        # VLOOKUP не поддерживается — данные в D1:E2 (вне зоны формулы A1),
        # чтобы не создавать циклическую ссылку на саму A1.
        ev = make_temp_eval({"S": {
            "A1": "=VLOOKUP(1,D1:E2,2,FALSE)",
            "D1": 1, "E1": "x", "D2": 2, "E2": "y",
        ***REMOVED******REMOVED***)
        with pytest.raises(FormulaError):
            ev.evaluate("S!A1")

    def test_circular_reference_raises(self, make_temp_eval):
        # A1 = A1 → циклическая ссылка → FormulaError (не RecursionError)
        ev = make_temp_eval({"S": {"A1": "=A1+1"***REMOVED******REMOVED***)
        with pytest.raises(FormulaError, match="[Cc***REMOVED***ircular"):
            ev.evaluate("S!A1")

    def test_unknown_sheet_raises(self, make_temp_eval):
        ev = make_temp_eval({"S": {"A1": 42***REMOVED******REMOVED***)
        with pytest.raises(FormulaError, match="Unknown sheet"):
            ev.evaluate("NOPE!A1")

    def test_bad_cell_address_raises(self, ev):
        with pytest.raises(FormulaError):
            ev.evaluate("order!X99Z")


# ---------------------------------------------------------------------------
# Согласованность: Excel-eval vs snapshot/Python recompute (BUG-005 proof)
# ---------------------------------------------------------------------------

class TestConsistency:
    def test_excel_eval_matches_snapshot(self, ev, snapshot, python_out):
        """Ключевой тест BUG-005: формулы из .xlsx (независимо вычисленные)
        совпадают с Python recompute и snapshot по всем SKU."""
        sku_rows = {4: "Молоко 3.2% 1л", 5: "Крупа гречневая 800г", 6: "Напиток газир. 1л"***REMOVED***
        for row, sku in sku_rows.items():
            excel_order = ev.evaluate(f"order!E{row***REMOVED***")
            excel_final = ev.evaluate(f"forecast!G{row***REMOVED***")

            py_order = python_out[sku***REMOVED***["order"***REMOVED***["order_qty"***REMOVED***
            py_final = python_out[sku***REMOVED***["forecast"***REMOVED***["final_forecast"***REMOVED***
            snap_order = snapshot["orders"***REMOVED***[sku***REMOVED***["order_qty"***REMOVED***
            snap_final = snapshot["forecasts"***REMOVED***[sku***REMOVED***["final"***REMOVED***

            assert excel_order == pytest.approx(py_order, abs=0.01)
            assert excel_order == pytest.approx(snap_order, abs=0.01)
            assert excel_final == pytest.approx(py_final, abs=0.01)
            assert excel_final == pytest.approx(snap_final, abs=0.01)
