"""
forecast.py — developer-role артефакт (Task 2, Mission B, promt 62).

Чистый Python-recompute прогноза без Excel engine, NO openpyxl expansion.
Named constants mirror `build_model_xlsx.py` (one source of truth) — никаких magic numbers.

Per Q1 (variant b): parity_check работает с этими outputs.
Per Q2: 3 категории (молочка / крупа / напиток).
Per Q3: 2 NON_OBVIOUS учтены в формулах (INCIDENT_2024_CORRECTION применён
только для категории `dairy`).
"""
from __future__ import annotations

import datetime
import json
***REMOVED***
from statistics import pstdev

from openpyxl import load_workbook

# === sys.path injection (run as direct script, not as -m module) ===
# Per thinker-with-files-gemini insight (ROADMAP-VV-001): direct script execution
# doesn't add cwd to sys.path; inject project root so `projects_17/...` resolves.
# Minimal-invasive option (d): NO __init__.py markers added.
import sys as _sys_for_path
***REMOVED*** as _Path_for_path
_sys_for_path.path.insert(
    0, str(_Path_for_path(__file__).resolve().parent.parent.parent)
)

# === ONE SOURCE OF TRUTH (imported from build_model_xlsx) ===
from projects_17.vkusvill_demo.build_model_xlsx import (
    BASE_DATE, WEEKS, CATEGORIES,
    SMA_WINDOW, SERVICE_LEVEL_Z, SHELF_CRITICAL_RATIO,
    INCIDENT_2024_CORRECTION, DEFAULT_STOCK, ORDER_BUFFER,
)

XLSX_PATH = Path("projects_17/vkusvill_demo/model_forecast.xlsx")
OUT_JSON = Path("projects_17/vkusvill_demo/forecast_python.json")


def load_history_from_xlsx(xlsx_path: Path) -> tuple[datetime.date, dict[str, list[int***REMOVED******REMOVED******REMOVED***:
    """Pure-Python read of history sheet — ТОЛЬКО raw sales_qty (no formula eval).

    Per Q1 variant (b): use data_only=True (read cached values), NEVER eval formulas.
    """
    wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
    hist = wb["history"***REMOVED***
    sku_sales: dict[str, list[int***REMOVED******REMOVED*** = {***REMOVED***
    cur_sku = None
    for row in hist.iter_rows(min_row=4, values_only=True):
        if row[2***REMOVED*** is None:  # blank separator line между категориями
            continue
        week, _date_str, sku, qty = row[0***REMOVED***, row[1***REMOVED***, row[2***REMOVED***, row[3***REMOVED***
        if cur_sku != sku:
            sku_sales[sku***REMOVED*** = [***REMOVED***
            cur_sku = sku
        sku_sales[sku***REMOVED***.append(int(qty))
    wb.close()
    return BASE_DATE, sku_sales


def compute_forecast(sku_sales: list[int***REMOVED***, cat: dict, base: datetime.date) -> dict:
    """Apply 4 принципа in pure Python (mirror build_model formulas).

    Returns dict with forecast components + final_forecast.
    """
    last4 = sku_sales[-SMA_WINDOW:***REMOVED***
    sma = sum(last4) / SMA_WINDOW
    sigma = pstdev(last4)
    wd_factor = cat["wd"***REMOVED***[(base.weekday() + WEEKS) % 7***REMOVED***
    safety = sigma * SERVICE_LEVEL_Z * (cat["lead_time"***REMOVED*** ** 0.5)
    shelf = 0.5 if cat["shelf_life"***REMOVED*** < cat["lead_time"***REMOVED*** * SHELF_CRITICAL_RATIO else 1.0
    final = ((sma * wd_factor) + safety) * shelf
    return {
        "sma": sma,
        "wd_factor": wd_factor,
        "safety_buffer": safety,
        "shelf_correction": shelf,
        "final_forecast": final,
    ***REMOVED***


def compute_order(forecast: dict, cat: dict, current_stock: int) -> dict:
    """Применяет правило `MAX(0, lead_time_demand - stock + buffer)`.

    NON_OBVIOUS_2: для категории `dairy` дополнительно применяется
    INCIDENT_2024_CORRECTION (legacy -8% rule post-инцидента 2024).
    """
    final = forecast["final_forecast"***REMOVED***
    order_qty = max(
        0,
        (final * cat["lead_time"***REMOVED***) - current_stock + (final * ORDER_BUFFER),
    )
    if cat["cat"***REMOVED*** == "dairy":
        order_qty *= INCIDENT_2024_CORRECTION
    return {"current_stock": current_stock, "order_qty": order_qty***REMOVED***


def main() -> None:
    if not XLSX_PATH.exists():
        raise SystemExit(f"ERR: {XLSX_PATH***REMOVED*** not found. Run build_model_xlsx.py first.")
    base, sku_sales = load_history_from_xlsx(XLSX_PATH)
    out: dict[str, dict***REMOVED*** = {***REMOVED***
    for cat in CATEGORIES:
        forecast = compute_forecast(sku_sales[cat["sku"***REMOVED******REMOVED***, cat, base)
        order = compute_order(forecast, cat, DEFAULT_STOCK[cat["sku"***REMOVED******REMOVED***)
        out[cat["sku"***REMOVED******REMOVED*** = {"forecast": forecast, "order": order***REMOVED***
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    total = sum(o["order"***REMOVED***["order_qty"***REMOVED*** for o in out.values())
    print(f"OK: {OUT_JSON***REMOVED***")
    print(f"Total order_qty (Python recompute) = {total:.2f***REMOVED***")


if __name__ == "__main__":
    main()
