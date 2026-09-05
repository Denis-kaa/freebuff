"""
core_02/xlsx_builder.py — минимальный Excel writer API (платформенный skill).

⚠️ Per PB-9 lesson (pyyaml recurrence): top-level `import openpyxl` даёт
loud `ModuleNotFoundError` при пропадании пакета из окружения — нет silent
fallback на bare openpyxl прячем в helpers.

Public API (per TestXxxBuilder contract):
    class Workbook:
        __init__(self) -> None
        sheet(self, name: str) -> None                 # switch/create
        cell(self, addr, *, value=..., formula=..., fmt=...) -> None   # XOR (value OR formula)
        save(self, path) -> Path                        # atomic: tempfile + os.replace
        load(cls, path, *, data_only=False) -> 'Workbook'   # classmethod
        cell_value(self, addr) -> Any
        cell_formula(self, addr) -> Optional[str***REMOVED***
        sheetnames(self) -> list[str***REMOVED***
        current_sheet (property) -> str

Per Task 0 scope: минимальный API (НЕ openpyxl wholesale), без define_name(),
без chart, без named ranges (минимальный набор для Skill сериализации).
"""
from __future__ import annotations

import os
import shutil
import tempfile
from copy import copy
***REMOVED***
from typing import Any, Dict, List, Optional, Union

import openpyxl
from openpyxl import Workbook as _XLWorkbook
from openpyxl.cell.cell import Cell

__all__ = ["Workbook"***REMOVED***


class Workbook:
    """Минимальный wrapper над openpyxl для сериализации .xlsx файлов."""

    def __init__(self) -> None:
        self._wb: _XLWorkbook = _XLWorkbook()
        self._current_name: str = self._wb.active.title

    def sheet(self, name: str) -> None:
        """Switch to existing sheet or create new one."""
        if name in self._wb.sheetnames:
            self._wb.active = self._wb[name***REMOVED***
        else:
            self._wb.create_sheet(title=name)
            self._wb.active = self._wb[name***REMOVED***
        self._current_name = name

    @property
    def current_sheet(self) -> str:
        return self._current_name

    def sheetnames(self) -> List[str***REMOVED***:
        return list(self._wb.sheetnames)

    def cell(
        self,
        addr: str,
        *,
        value: Any = None,
        formula: Optional[str***REMOVED*** = None,
        fmt: Optional[Dict[str, Any***REMOVED******REMOVED*** = None,
    ) -> None:
        """Set cell content. value XOR formula (не оба, не ни одного).

        Args:
            addr: cell address like 'A1', 'B17'. str only.
            value: literal value (int, str, float, bool, None).
            formula: Excel formula string starting with '='. Optional.
            fmt: format dict, optional keys 'bold', 'fill', 'align'.
        """
        if value is not None and formula is not None:
            raise ValueError(f"cell({addr***REMOVED***): value XOR formula constraint violated")
        if value is None and formula is None:
            raise ValueError(f"cell({addr***REMOVED***): must pass either value or formula")

        c: Cell = self._wb.active[addr***REMOVED***
        if formula is not None:
            # Per thinker-with-files-gemini insight (ROADMAP-VV-001): openpyxl
            # tracks formula vs value via `=` prefix. Without it, `data_only=True`
            # on load won't return None for unevaluated formulas. Auto-prepend.
            c.value = formula if formula.startswith("=") else f"={formula***REMOVED***"
        else:
            c.value = value

        if fmt:
            if fmt.get("bold"):
                # Per openpyxl 3.1+ deprecation: use stdlib copy + attribute set.
                new_font = copy(c.font)
                new_font.bold = True
                c.font = new_font
            if fmt.get("fill"):
                c.fill = openpyxl.styles.PatternFill(
                    start_color=fmt["fill"***REMOVED***, end_color=fmt["fill"***REMOVED***, fill_type="solid"
                )
            if fmt.get("align"):
                c.alignment = openpyxl.styles.Alignment(horizontal=fmt["align"***REMOVED***)

    def cell_value(self, addr: str) -> Any:
        """Read cell value (literal, не formula)."""
        return self._wb.active[addr***REMOVED***.value

    def cell_formula(self, addr: str) -> Optional[str***REMOVED***:
        """Read cell formula (if any)."""
        v = self._wb.active[addr***REMOVED***.value
        if isinstance(v, str) and v.startswith("="):
            return v
        return None

    def save(self, path: Union[str, Path***REMOVED***) -> Path:
        """Save atomic via tempfile in SAME DIR + os.replace (PB-7 mirror: fuseblk safe).

        Per PB-7 (cross-device save on fuseblk): tempfile.mkstemp в /tmp нельзя
        переносить через os.replace на /storage (fuseblk) — OSError [Errno 18***REMOVED***
        Cross-device link. Fix: tempfile.mkstemp с dir=path.parent, чтобы
        temp и destination были на одном mount.

        Returns:
            Path to saved file.
        """
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        fd, tmp_path = tempfile.mkstemp(
            suffix=".xlsx", prefix=".xlsx_builder_", dir=str(path.parent)
        )
        os.close(fd)
        try:
            self._wb.save(tmp_path)
            os.replace(tmp_path, path)
        except Exception:
            if os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
            raise
        return path

    @classmethod
    def load(cls, path: Union[str, Path***REMOVED***, *, data_only: bool = False) -> "Workbook":
        """Load existing .xlsx. data_only=True — read cached values, NO formula eval."""
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"xlsx not found: {path***REMOVED***")
        wb = cls.__new__(cls)
        wb._wb = openpyxl.load_workbook(str(path), data_only=data_only)
        wb._current_name = wb._wb.active.title
        return wb
