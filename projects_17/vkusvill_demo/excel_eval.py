"""
excel_eval.py — независимый Excel-formula evaluator (BUG-005 fix, promt 64 audit).

Проблема (BUG-005): parity_check.py сравнивал два Python-источника
(model_snapshot.json от build_model_xlsx.py + forecast_python.json от forecast.py),
а реальные Excel-формулы в model_forecast.xlsx никогда не вычислялись.
«Excel-vs-Python эквивалентность» была заявлена, но не доказана.

Решение: этот модуль читает ФОРМУЛЫ прямо из .xlsx (data_only=False)
и вычисляет их независимым парсером. Это настоящий Excel-eval путь:
формулы из файла интерпретируются заново (AVERAGE, STDEV.P, SQRT, IF,
MAX, SUM, арифметика, cross-sheet и range ссылки) — без Excel/LibreOffice
и без повторного использования Python-логики build_model/forecast.

Почему не pycel / formulas / LibreOffice (Termux, Python 3.14, ARM64):
- pycel установился, но падает на py3.14: `module 'ast' has no attribute 'Str'`
  (ast.Str удалён в 3.12+, pycel использует устаревший API).
- formulas требует numpy/pandas — не ставится за разумное время на ARM64 Termux.
- LibreOffice на Termux недоступен (soffice NOT_FOUND).
- Собственный evaluator покрывает ровно подмножество формул demo (см. SUPPORTED)
  и не добавляет зависимостей (только openpyxl, уже в requirements.txt).

SUPPORTED formula subset (ровно то, что встречается в model_forecast.xlsx):
  - литералы: числа (int/float)
  - ссылки: A1, 'Sheet'!A1, диапазоны A1:B2, 'Sheet'!A1:B2
  - функции: AVERAGE(range), STDEV.P(range), SQRT(x), IF(cond,a,b),
    MAX(a,b,...), SUM(range|a,b,...)
  - операторы: + - * / ( ), сравнения < > <= >= = <>
  - вложенные формулы (ячейка со ссылкой на другую формулу) — резолвятся рекурсивно
"""
from __future__ import annotations

import re
from statistics import pstdev
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Tokenizer
# ---------------------------------------------------------------------------

_TOKEN_RE = re.compile(
    r"""
    (?P<WS>\s+)
  | (?P<QUOTED>'[^')*')
  | (?P<NUM>\d+(?:\.\d+)?)
  | (?P<NAME>[A-Za-z_.][A-Za-z0-9_.]*)
  | (?P<OP><=|>=|<>|[<>=+\-*/])
  | (?P<PUNCT>[(),:!])
    """,
    re.VERBOSE,
)


class FormulaError(Exception):
    pass


def tokenize(formula: str) -> list[tuple[str, str]]:
    tokens: list[tuple[str, str]] = []
    pos = 0
    for m in _TOKEN_RE.finditer(formula):
        if m.start() != pos:
            raise FormulaError(f"Unexpected char {formula[pos]!r} in {formula!r}")
        pos = m.end()
        kind = m.lastgroup
        if kind == "WS":
            continue
        tokens.append((kind, m.group()))
    if pos != len(formula):
        raise FormulaError(f"Trailing chars in {formula!r}: {formula[pos:]!r}")
    tokens.append(("END", ""))
    return tokens


# ---------------------------------------------------------------------------
# Evaluator
# ---------------------------------------------------------------------------

class ExcelEval:
    """Ленивый evaluator формул xlsx.

    Загружает workbook один раз (data_only=False → формулы доступны как строки),
    вычисляет значения по требованию с кэшем. Поддерживает cross-sheet ссылки
    ('history'!D4:D15), вложенные формулы и подмножество функций из SUPPORTED.
    """

    def __init__(self, xlsx_path: str | Path) -> None:
        self.wb = load_workbook(str(xlsx_path), data_only=False)
        # sheet -> {(col_letter, row_number): raw_value}
        self._raw: dict[str, dict[tuple[str, int], Any]] = {}
        for name in self.wb.sheetnames:
            ws = self.wb[name]
            sheet: dict[tuple[str, int], Any] = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        sheet[(cell.column_letter, cell.row)] = cell.value
            self._raw[name] = sheet
        self._cache: dict[tuple[str, str, int], Any] = {}
        self._evaluating: set[tuple[str, str, int]] = set()

    # -- helpers ------------------------------------------------------------

    def _cell(self, sheet: str, col: str, row: int) -> Any:
        """Возвращает вычисленное значение ячейки (рекурсивно для формул).

        С детекцией циклических ссылок: если ячейка уже в процессе вычисления —
        raise FormulaError (иначе бесконечная рекурсия → RecursionError).
        """
        key = (sheet, col, row)
        if key in self._cache:
            return self._cache[key]
        if key in self._evaluating:
            raise FormulaError(
                f"Circular reference detected: {sheet}!{col}{row}"
            )
        if sheet not in self._raw:
            raise FormulaError(f"Unknown sheet {sheet!r}")
        raw = self._raw[sheet].get((col, row))
        if raw is None:
            self._cache[key] = None
            return None
        if isinstance(raw, str) and raw.startswith("="):
            self._evaluating.add(key)
            try:
                self._cache[key] = self._eval_formula(sheet, raw[1:])
            except FormulaError as exc:
                raise FormulaError(f"{sheet}!{col}{row}: {exc}") from exc
            finally:
                self._evaluating.discard(key)
            return self._cache[key]
        self._cache[key] = raw
        return raw

    def _eval_formula(self, sheet: str, body: str) -> Any:
        tokens = tokenize(body)
        parser = _Parser(self, sheet, tokens)
        val = parser.parse_expression()
        if parser.peek()[0] != "END":
            raise FormulaError(f"Trailing tokens after expression: {body!r}")
        return val

    def evaluate(self, address: str) -> Any:
        """address вида 'Sheet!A1' или 'A1' (текущий первый лист)."""
        if "!" in address:
            sheet, cell = address.split("!", 1)
            sheet = sheet.strip("'")
        else:
            sheet = self.wb.sheetnames[0]
            cell = address
        m = re.fullmatch(r"([A-Za-z)+)(\d+)", cell.strip())
        if not m:
            raise FormulaError(f"Bad cell address {address!r}")
        return self._cell(sheet, m.group(1).upper(), int(m.group(2)))

    def find_formula_cell(
        self, sheet: str, col: str, prefix: str = "=SUM("
    ) -> tuple[str, int] | None:
        """Найти первую ячейку в колонке, чья формула начинается с prefix.

        Public helper (used by parity_check для поиска TOTAL-строки), чтобы не
        лезть в приватный `_raw` из другого модуля.
        """
        if sheet not in self._raw:
            return None
        for (c, row), raw in self._raw[sheet].items():
            if (
                c == col.upper()
                and isinstance(raw, str)
                and raw.startswith(prefix)
            ):
                return (c, row)
        return None


# ---------------------------------------------------------------------------
# Recursive-descent parser
# ---------------------------------------------------------------------------

class _Parser:
    def __init__(self, ev: ExcelEval, sheet: str, tokens: list[tuple[str, str]]) -> None:
        self.ev = ev
        self.sheet = sheet
        self.toks = tokens
        self.i = 0

    def peek(self) -> tuple[str, str]:
        return self.toks[self.i]

    def next(self) -> tuple[str, str]:
        tok = self.toks[self.i]
        self.i += 1
        return tok

    # precedence: comparison < additive < multiplicative < unary < primary
    def parse_expression(self) -> Any:
        left = self.parse_additive()
        kind, op = self.peek()
        if kind == "OP" and op in ("<", ">", "<=", ">=", "=", "<>"):
            self.next()
            right = self.parse_additive()
            return _compare(op, left, right)
        return left

    def parse_additive(self) -> Any:
        val = self.parse_multiplicative()
        while True:
            kind, op = self.peek()
            if kind == "OP" and op in ("+", "-"):
                self.next()
                rhs = self.parse_multiplicative()
                val = val + rhs if op == "+" else val - rhs
            else:
                return val

    def parse_multiplicative(self) -> Any:
        val = self.parse_unary()
        while True:
            kind, op = self.peek()
            if kind == "OP" and op in ("*", "/"):
                self.next()
                rhs = self.parse_unary()
                val = val * rhs if op == "*" else val / rhs
            else:
                return val

    def parse_unary(self) -> Any:
        kind, op = self.peek()
        if kind == "OP" and op == "-":
            self.next()
            return -self.parse_unary()
        return self.parse_primary()

    def parse_primary(self) -> Any:
        kind, tok = self.next()
        if kind == "NUM":
            return float(tok)
        if kind == "PUNCT" and tok == "(":
            val = self.parse_expression()
            k, _ = self.next()
            if k != "PUNCT" or _ != ")":
                raise FormulaError("Expected ')'")
            return val
        if kind == "NAME":
            # функция или ссылка
            nk, nxt = self.peek()
            if nk == "PUNCT" and nxt == "(":
                self.next()  # consume '('
                args: list[Any] = []
                while True:
                    ak, anxt = self.peek()
                    if ak == "PUNCT" and anxt == ")":
                        self.next()
                        break
                    # range-аргумент: 'Sheet'!A1:B2 или A1:B2
                    if self._looks_like_range():
                        args.append(self._parse_range_arg())
                    else:
                        args.append(self.parse_expression())
                    ak2, anxt2 = self.peek()
                    if ak2 == "PUNCT" and anxt2 == ",":
                        self.next()
                    elif ak2 == "PUNCT" and anxt2 == ")":
                        self.next()
                        break
                    else:
                        raise FormulaError("Expected ',' or ')' in function args")
                return _call(self.ev, tok.upper(), args)
            # ссылка на ячейку/диапазон
            return self._parse_cellref(tok)
        if kind == "QUOTED":
            # 'Sheet'!A1  или  'Sheet'!A1:B2
            sheet = tok.strip("'")
            nk, nxt = self.next()
            if nk != "PUNCT" or nxt != "!":
                raise FormulaError(f"Expected '!' after sheet name {tok!r}")
            nk2, cell_or_range = self.next()
            if nk2 != "NAME":
                raise FormulaError(f"Expected cell ref after '!', got {cell_or_range!r}")
            return self._parse_cellref(cell_or_range, sheet=sheet)
        raise FormulaError(f"Unexpected token {kind}={tok!r}")

    def _looks_like_range(self) -> bool:
        """A1:B2  или  'Sheet'!A1:B2 — начинается как ссылка и содержит ':'."""
        # Кэшируем позицию; проверяем до двух токенов вперёд
        j = self.i
        save = self.i
        try:
            sheet = None
            kind, tok = self.toks[j]; j += 1
            if kind == "QUOTED":
                sheet = tok
                k1, t1 = self.toks[j]; j += 1
                if not (k1 == "PUNCT" and t1 == "!"):
                    return False
                k2, t2 = self.toks[j]; j += 1
                if k2 != "NAME":
                    return False
                cell1 = t2
            elif kind == "NAME":
                cell1 = tok
            else:
                return False
            k3, t3 = self.toks[j]; j += 1
            if k3 == "PUNCT" and t3 == ":":
                return True
            return False
        finally:
            self.i = save

    def _parse_range_arg(self) -> list[Any]:
        """Парсит 'Sheet'!A1:B2 или A1:B2 → СПИСОК ЗНАЧЕНИЙ ячеек диапазона."""
        sheet = self.sheet
        kind, tok = self.next()
        if kind == "QUOTED":
            sheet = tok.strip("'")
            k, b = self.next()
            if not (k == "PUNCT" and b == "!"):
                raise FormulaError("Expected '!'")
            kind, tok = self.next()
            if kind != "NAME":
                raise FormulaError("Expected cell ref")
        cell1 = tok
        k, colon = self.next()
        if not (k == "PUNCT" and colon == ":"):
            raise FormulaError("Expected ':' in range")
        kind, cell2 = self.next()
        if kind != "NAME":
            raise FormulaError("Expected second cell ref in range")
        sheet, c1, r1, c2, r2 = _split_range(sheet, cell1, cell2)
        return _range_values(self.ev, sheet, c1, r1, c2, r2)

    def _parse_cellref(self, name: str, sheet: str | None = None) -> Any:
        """Одиночная ссылка A1 → значение.

        NOTE: диапазоны (A1:B2) НЕ обрабатываются здесь — токенизатор разделяет
        ':' как отдельный PUNCT, поэтому NAME-токен никогда не содержит ':'. Диапазоны
        приходят только через `_parse_range_arg` (аргументы функций AVERAGE/SUM/...).
        """
        sheet = sheet or self.sheet
        m = re.fullmatch(r"([A-Za-z)+)(\d+)", name)
        if not m:
            raise FormulaError(f"Bad reference {name!r}")
        return self.ev._cell(sheet, m.group(1).upper(), int(m.group(2)))


# ---------------------------------------------------------------------------
# Function dispatch
# ---------------------------------------------------------------------------

def _split_range(sheet: str, cell1: str, cell2: str) -> tuple[str, str, int, str, int]:
    m1 = re.fullmatch(r"([A-Za-z)+)(\d+)", cell1)
    m2 = re.fullmatch(r"([A-Za-z)+)(\d+)", cell2)
    if not m1 or not m2:
        raise FormulaError(f"Bad range {cell1}:{cell2}")
    return sheet, m1.group(1).upper(), int(m1.group(2)), m2.group(1).upper(), int(m2.group(2))


def _range_values(ev: ExcelEval, sheet: str, c1: str, r1: int, c2: str, r2: int) -> list[Any]:
    cols = _expand_cols(c1, c2)
    values: list[Any] = []
    for col in cols:
        for row in range(min(r1, r2), max(r1, r2) + 1):
            v = ev._cell(sheet, col, row)
            if v is not None:
                values.append(v)
    return values


def _expand_cols(c1: str, c2: str) -> list[str]:
    def col_num(c: str) -> int:
        n = 0
        for ch in c:
            n = n * 26 + (ord(ch) - ord("A") + 1)
        return n

    def col_str(n: int) -> str:
        s = ""
        while n:
            n, rem = divmod(n - 1, 26)
            s = chr(ord("A") + rem) + s
        return s

    a, b = col_num(c1), col_num(c2)
    lo, hi = min(a, b), max(a, b)
    return [col_str(n) for n in range(lo, hi + 1)]


def _compare(op: str, a: Any, b: Any) -> bool:
    if op == "<":
        return a < b
    if op == ">":
        return a > b
    if op == "<=":
        return a <= b
    if op == ">=":
        return a >= b
    if op == "=":
        return a == b
    if op == "<>":
        return a != b
    raise FormulaError(f"Unknown comparison {op!r}")


def _call(ev: ExcelEval, name: str, args: list[Any]) -> Any:
    name = name.upper()
    if name == "AVERAGE":
        vals = _flatten(args)
        return sum(vals) / len(vals)
    if name == "STDEV.P":
        vals = _flatten(args)
        return pstdev(vals)
    if name == "SQRT":
        return args[0] ** 0.5
    if name == "IF":
        if len(args) != 3:
            raise FormulaError(f"IF expects 3 arguments, got {len(args)}")
        cond, then_v, else_v = args[0], args[1], args[2]
        return then_v if cond else else_v
    if name == "MAX":
        vals = _flatten(args)
        return max(vals)
    if name == "SUM":
        vals = _flatten(args)
        return sum(vals)
    raise FormulaError(f"Unsupported function {name}")


def _flatten(args: list[Any]) -> list[Any]:
    out: list[Any] = []
    for a in args:
        if isinstance(a, list):
            out.extend(a)
        else:
            out.append(a)
    return out


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    import json
    import sys

    xlsx = Path("projects_17/vkusvill_demo/model_forecast.xlsx")
    out_json = Path("projects_17/vkusvill_demo/excel_eval.json")
    if not xlsx.exists():
        print(f"ERR: {xlsx} not found")
        return 2
    ev = ExcelEval(xlsx)

    result: dict[str, Any] = {"orders": {}, "forecasts": {}}
    # order sheet E4:E6 = order_qty; TOTAL row = SUM(E4:E6) — найти динамически
    total_addr = None
    hit = ev.find_formula_cell("order", "E", prefix="=SUM(")
    if hit:
        total_addr = f"order!{hit[0]}{hit[1]}"
    # forecast G4:G6 = final_forecast, order E4:E6 = order_qty
    for row in range(4, 7):
        result["orders"].setdefault("order_qty", {})[f"E{row}"] = ev.evaluate(f"order!E{row}")
        result["forecasts"].setdefault("final_forecast", {})[f"G{row}"] = ev.evaluate(f"forecast!G{row}")
    result["orders"]["total_addr"] = total_addr
    result["orders"]["total"] = ev.evaluate(total_addr) if total_addr else None

    out_json.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK: {out_json}")
    for k, v in result["orders"]["order_qty"].items():
        print(f"order!{k} (Excel-eval) = {v:.4f}")
    if result["orders"]["total"] is not None:
        print(f"{result['orders']['total_addr']} total (Excel-eval) = {result['orders']['total']:.4f}")
    for k, v in result["forecasts"]["final_forecast"].items():
        print(f"forecast!{k} (Excel-eval) = {v:.4f}")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
