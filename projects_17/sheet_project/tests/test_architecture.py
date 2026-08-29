"""Архитектурные тесты (invariant-защита) — `tests/test_architecture.py`.

Защищают ключевые инварианты D2 (architecture.md §2.4/§5/§7,
contracts.yaml `generator.invariants` + `dependency_direction`):

1. «Ядро не знает шаблона» — `generator/*` не импортирует `config.project_dashboard` (R2).
2. «openpyxl изолирован» — openpyxl только в `generator/*` и `validator/*` (R5).
3. «Смена CONFIG без правки ядра» — два разных CONFIG → два структурно разных XLSX (это и есть D2).
4. «Направление зависимостей» — `validator/*`, `config/*`, `data/*`, `styles/*` не импортируют `generator/*` (`dependency_direction.forbidden`).

(1), (2) и (4) — статические (AST), активны сразу; (3) — поведенческий, включается
на этапе 6 (реализация `generator/*`) — до этого помечен skip.
"""

from __future__ import annotations

import ast
***REMOVED***

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _py_files(directory: Path) -> list[Path***REMOVED***:
    if not directory.exists():
        return [***REMOVED***
    return sorted(directory.rglob("*.py"))


def _imports_module(path: Path, module: str) -> bool:
    """True, если исходник импортирует `module` в любой форме.

    Покрывает `import X`, `from X import ...` и `from parent import child`
    (где полный путь `parent.child` == module, напр. `from config import project_dashboard`).
    """
    tree = ast.parse(path.read_text(encoding="utf-8"))
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            if any(alias.name == module for alias in node.names):
                return True
        elif isinstance(node, ast.ImportFrom):
            if node.module == module:
                return True
            if node.module:
                for alias in node.names:
                    if f"{node.module***REMOVED***.{alias.name***REMOVED***" == module:
                        return True
    return False


def _imports_under(path: Path, package: str) -> bool:
    """True, если исходник импортирует `package` или любой подмодуль `package.*`.

    Покрывает `import generator`, `import generator.workbook`,
    `from generator import x`, `from generator.workbook import y`.
    """
    tree = ast.parse(path.read_text(encoding="utf-8"))
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                if alias.name == package or alias.name.startswith(package + "."):
                    return True
        elif isinstance(node, ast.ImportFrom):
            if node.module and (node.module == package or node.module.startswith(package + ".")):
                return True
    return False


# ─────────────────────────────────────────────────────────────────────────────
# 1. Ядро не знает шаблона (R2)
# ─────────────────────────────────────────────────────────────────────────────

def test_generator_does_not_import_project_dashboard():
    files = _py_files(PROJECT_ROOT / "generator")
    if not files:
        pytest.skip("generator/ ещё не реализован (этап 6)")
    offenders = [str(f) for f in files if _imports_module(f, "config.project_dashboard")***REMOVED***
    assert not offenders, f"Ядро импортирует шаблон (нарушение R2): {offenders***REMOVED***"


# ─────────────────────────────────────────────────────────────────────────────
# 2. openpyxl изолирован (R5)
# ─────────────────────────────────────────────────────────────────────────────

def test_openpyxl_isolation():
    offenders = [***REMOVED***
    for dirname in ("config", "data", "styles"):
        for f in _py_files(PROJECT_ROOT / dirname):
            if _imports_under(f, "openpyxl"):
                offenders.append(str(f))
    assert not offenders, f"openpyxl вне generator/validator (нарушение R5): {offenders***REMOVED***"


def test_dependency_direction_no_upstream_import():
    """Нижние слои не импортируют generator/* (dependency_direction.forbidden)."""
    offenders = [***REMOVED***
    for dirname in ("validator", "config", "data", "styles"):
        for f in _py_files(PROJECT_ROOT / dirname):
            if _imports_under(f, "generator"):
                offenders.append(str(f))
    assert not offenders, f"Запрещённый импорт generator/* из нижнего слоя: {offenders***REMOVED***"


# ─────────────────────────────────────────────────────────────────────────────
# 3. Смена CONFIG без правки ядра (D2)
# ─────────────────────────────────────────────────────────────────────────────

@pytest.mark.skip(reason="GENERATOR не реализован (этап 6): включить после generator/workbook.py")
def test_config_swap_changes_output_without_core_edit(tmp_path, monkeypatch):
    # TODO(этап 6) зафиксировать ДО снятия skip:
    #  1) theme — по контракту обязателен (Theme, не None): styles/theme.py даст load_theme()
    #     на этапе 5; здесь заменить None на реальный Theme.
    #  2) OUTPUT_DIR — contracts.yaml `generator.writes` хардкодит output/*.xlsx; на этапе 6
    #     согласовать: generate() либо чтит OUTPUT_DIR, либо тест пишет во временный output/
    #     через явный параметр (не env).
    from config.schema import ArtifactStatus, DataSource, Field, FieldType, Sheet, Workbook
    from data.models import Project
    from generator.workbook import generate

    def make_config(sheet_name: str, columns: list[str***REMOVED***) -> Workbook:
        sheet = Sheet(
            name=sheet_name,
            columns=[Field(name=c, type=FieldType.TEXT) for c in columns***REMOVED***,
            data_source=DataSource(source="rows", field_map={c: c for c in columns***REMOVED***),
        )
        return Workbook(name="cfg", template_id="cfg", template_version="1.0", sheets=[sheet***REMOVED***)

    data = {
        "rows": [
            Project(id="1", name="Альфа", status="in_progress"),
            Project(id="2", name="Бета", status="done"),
        ***REMOVED***
    ***REMOVED***
    monkeypatch.setenv("OUTPUT_DIR", str(tmp_path))  # не писать в реальный output/

    artifact_a = generate(make_config("ProjectsA", ["id", "name"***REMOVED***), data, theme=None, output_name="a.xlsx")
    artifact_b = generate(make_config("ProjectsB", ["name", "status"***REMOVED***), data, theme=None, output_name="b.xlsx")

    assert artifact_a.status is ArtifactStatus.READY
    assert artifact_b.status is ArtifactStatus.READY

    import openpyxl

    wb_a = openpyxl.load_workbook(artifact_a.path)
    wb_b = openpyxl.load_workbook(artifact_b.path)
    # Смена CONFIG (имя листа + набор колонок) дала другой XLSX — без правки ядра.
    assert wb_a.sheetnames != wb_b.sheetnames
    header_a = [c.value for c in wb_a.active[1***REMOVED******REMOVED***
    header_b = [c.value for c in wb_b.active[1***REMOVED******REMOVED***
    assert header_a != header_b  # наборы колонок различаются
